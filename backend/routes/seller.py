"""
Seller routes - statistics and seller-specific utilities
"""
from flask import Blueprint, jsonify, request, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from sqlalchemy import func, extract, case
from datetime import datetime, timedelta
import io
import csv

from models import db
from models.user import User
from models.tour import Tour
from models.booking import Booking
from models.seller_tour_guide import SellerTourGuide

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not available. Excel export will be disabled.")

seller_bp = Blueprint('seller', __name__, url_prefix='/api/seller')


@seller_bp.route('/stats', methods=['GET'])
@jwt_required()
def seller_stats():
    """Return aggregated stats for the authenticated seller (or admin).

    Response:
      - total_tours
      - bookings_count
      - pending_bookings_count
      - confirmed_bookings_count
      - cancelled_bookings_count
      - income_sum
      - confirmed_income_sum (income from confirmed bookings only)
      - average_rating
    """
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        # Admins can view global aggregates
        if user.role == 'admin':
            total_tours = db.session.query(func.count(Tour.id)).scalar() or 0
            bookings_count = db.session.query(func.count(Booking.id)).scalar() or 0
            pending_bookings_count = db.session.query(func.count(Booking.id)).filter(Booking.status == 'pending').scalar() or 0
            confirmed_bookings_count = db.session.query(func.count(Booking.id)).filter(Booking.status == 'confirmed').scalar() or 0
            cancelled_bookings_count = db.session.query(func.count(Booking.id)).filter(Booking.status == 'cancelled').scalar() or 0
            income_sum = db.session.query(func.coalesce(func.sum(Booking.total_price), 0)).scalar() or 0
            confirmed_income_sum = db.session.query(func.coalesce(func.sum(Booking.total_price), 0)).filter(Booking.status == 'confirmed').scalar() or 0
            average_rating = db.session.query(func.coalesce(func.avg(Tour.rating), 0)).scalar() or 0
        else:
            total_tours = db.session.query(func.count(Tour.id)).filter(Tour.seller_id == current_user_id).scalar() or 0
            
            # Base query for seller's bookings
            seller_bookings = db.session.query(Booking).join(Tour, Booking.tour_id == Tour.id).filter(Tour.seller_id == current_user_id)
            
            bookings_count = seller_bookings.count() or 0
            pending_bookings_count = seller_bookings.filter(Booking.status == 'pending').count() or 0
            confirmed_bookings_count = seller_bookings.filter(Booking.status == 'confirmed').count() or 0
            cancelled_bookings_count = seller_bookings.filter(Booking.status == 'cancelled').count() or 0
            
            income_sum = (
                db.session.query(func.coalesce(func.sum(Booking.total_price), 0))
                .join(Tour, Booking.tour_id == Tour.id)
                .filter(Tour.seller_id == current_user_id)
                .scalar()
                or 0
            )
            confirmed_income_sum = (
                db.session.query(func.coalesce(func.sum(Booking.total_price), 0))
                .join(Tour, Booking.tour_id == Tour.id)
                .filter(Tour.seller_id == current_user_id)
                .filter(Booking.status == 'confirmed')
                .scalar()
                or 0
            )
            average_rating = (
                db.session.query(func.coalesce(func.avg(Tour.rating), 0))
                .filter(Tour.seller_id == current_user_id)
                .scalar()
                or 0
            )

        return jsonify({
            'total_tours': int(total_tours),
            'bookings_count': int(bookings_count),
            'pending_bookings_count': int(pending_bookings_count),
            'confirmed_bookings_count': int(confirmed_bookings_count),
            'cancelled_bookings_count': int(cancelled_bookings_count),
            'income_sum': float(income_sum),
            'confirmed_income_sum': float(confirmed_income_sum),
            'average_rating': float(average_rating)
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error computing stats: {str(e)}'}), 500


@seller_bp.route('/assign-all-tours', methods=['POST'])
@jwt_required()
def assign_all_tours():
    """Assign all tours to the current seller (for development/testing)"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if user.role != 'seller' and user.role != 'admin':
            return jsonify({'error': 'Only sellers and admins can assign tours'}), 403
        
        # Get all tours
        all_tours = Tour.query.all()
        updated_count = 0
        
        for tour in all_tours:
            if tour.seller_id != current_user_id:
                tour.seller_id = current_user_id
                updated_count += 1
        
        if updated_count > 0:
            db.session.commit()
            return jsonify({
                'success': True,
                'message': f'Assigned {updated_count} tours to {user.username}',
                'updated_count': updated_count,
                'total_tours': len(all_tours)
            }), 200
        else:
            return jsonify({
                'success': True,
                'message': 'All tours are already assigned to you',
                'updated_count': 0,
                'total_tours': len(all_tours)
            }), 200
    
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error assigning tours: {str(e)}'}), 500


@seller_bp.route('/tour-guides', methods=['GET'])
@jwt_required()
def get_seller_tour_guides():
    """Get list of tour guides for the authenticated seller"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if user.role != 'seller' and user.role != 'admin':
            return jsonify({'error': 'Only sellers and admins can view tour guides'}), 403
        
        # Get all tour guides for this seller
        seller_tour_guides = SellerTourGuide.query.filter_by(seller_id=current_user_id).all()
        
        tour_guides = [stg.to_dict(include_guide=True)['tour_guide'] for stg in seller_tour_guides]
        
        return jsonify({
            'success': True,
            'tour_guides': tour_guides,
            'total': len(tour_guides)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error fetching tour guides: {str(e)}'}), 500


@seller_bp.route('/tour-guides', methods=['POST'])
@jwt_required()
def add_seller_tour_guide():
    """Add a tour guide to seller's list"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if user.role != 'seller' and user.role != 'admin':
            return jsonify({'error': 'Only sellers and admins can add tour guides'}), 403
        
        data = request.get_json()
        if not data or 'tour_guide_id' not in data:
            return jsonify({'error': 'tour_guide_id is required'}), 400
        
        tour_guide_id = int(data['tour_guide_id'])
        
        # Check if tour guide exists and has correct role
        tour_guide = User.query.get(tour_guide_id)
        if not tour_guide:
            return jsonify({'error': 'Tour guide not found'}), 404
        
        if tour_guide.role != 'tour_guide':
            return jsonify({'error': 'Selected user is not a tour guide'}), 400
        
        # Check if already exists
        existing = SellerTourGuide.query.filter_by(
            seller_id=current_user_id,
            tour_guide_id=tour_guide_id
        ).first()
        
        if existing:
            return jsonify({
                'success': True,
                'message': 'Tour guide already in your list',
                'tour_guide': existing.to_dict(include_guide=True)['tour_guide']
            }), 200
        
        # Create new relationship
        seller_tour_guide = SellerTourGuide(
            seller_id=current_user_id,
            tour_guide_id=tour_guide_id
        )
        
        db.session.add(seller_tour_guide)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Tour guide added successfully',
            'tour_guide': seller_tour_guide.to_dict(include_guide=True)['tour_guide']
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error adding tour guide: {str(e)}'}), 500


@seller_bp.route('/tour-guides/<int:tour_guide_id>', methods=['DELETE'])
@jwt_required()
def remove_seller_tour_guide(tour_guide_id):
    """Remove a tour guide from seller's list"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        if user.role != 'seller' and user.role != 'admin':
            return jsonify({'error': 'Only sellers and admins can remove tour guides'}), 403
        
        # Find and delete the relationship
        seller_tour_guide = SellerTourGuide.query.filter_by(
            seller_id=current_user_id,
            tour_guide_id=tour_guide_id
        ).first()
        
        if not seller_tour_guide:
            return jsonify({'error': 'Tour guide not found in your list'}), 404
        
        db.session.delete(seller_tour_guide)
        db.session.commit()
        
        return jsonify({
            'success': True,
            'message': 'Tour guide removed successfully'
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error removing tour guide: {str(e)}'}), 500


@seller_bp.route('/revenue-stats', methods=['GET'])
@jwt_required()
def revenue_stats():
    """Get detailed revenue statistics by day/month/year"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        period = request.args.get('period', 'month')  # day, month, year
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')

        # Base query for seller's bookings
        if user.role == 'admin':
            base_query = db.session.query(Booking)
        else:
            base_query = db.session.query(Booking).join(Tour, Booking.tour_id == Tour.id).filter(Tour.seller_id == current_user_id)

        # Apply date filters if provided
        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                base_query = base_query.filter(Booking.created_at >= start_dt)
            except ValueError:
                pass

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                end_dt = end_dt + timedelta(days=1)  # Include the entire end date
                base_query = base_query.filter(Booking.created_at < end_dt)
            except ValueError:
                pass

        stats = []
        
        if period == 'day':
            # Group by day
            results = (
                base_query
                .with_entities(
                    func.date(Booking.created_at).label('date'),
                    func.count(Booking.id).label('count'),
                    func.sum(case((Booking.status == 'confirmed', Booking.total_price), else_=0)).label('confirmed_revenue'),
                    func.sum(Booking.total_price).label('total_revenue')
                )
                .group_by(func.date(Booking.created_at))
                .order_by(func.date(Booking.created_at).desc())
                .all()
            )
            
            for row in results:
                stats.append({
                    'period': row.date.strftime('%Y-%m-%d'),
                    'label': row.date.strftime('%d/%m/%Y'),
                    'bookings_count': int(row.count),
                    'confirmed_revenue': float(row.confirmed_revenue or 0),
                    'total_revenue': float(row.total_revenue or 0)
                })
        
        elif period == 'month':
            # Group by month
            results = (
                base_query
                .with_entities(
                    extract('year', Booking.created_at).label('year'),
                    extract('month', Booking.created_at).label('month'),
                    func.count(Booking.id).label('count'),
                    func.sum(case((Booking.status == 'confirmed', Booking.total_price), else_=0)).label('confirmed_revenue'),
                    func.sum(Booking.total_price).label('total_revenue')
                )
                .group_by(extract('year', Booking.created_at), extract('month', Booking.created_at))
                .order_by(extract('year', Booking.created_at).desc(), extract('month', Booking.created_at).desc())
                .all()
            )
            
            for row in results:
                stats.append({
                    'period': f"{int(row.year)}-{int(row.month):02d}",
                    'label': f"Tháng {int(row.month)}/{int(row.year)}",
                    'bookings_count': int(row.count),
                    'confirmed_revenue': float(row.confirmed_revenue or 0),
                    'total_revenue': float(row.total_revenue or 0)
                })
        
        elif period == 'year':
            # Group by year
            results = (
                base_query
                .with_entities(
                    extract('year', Booking.created_at).label('year'),
                    func.count(Booking.id).label('count'),
                    func.sum(case((Booking.status == 'confirmed', Booking.total_price), else_=0)).label('confirmed_revenue'),
                    func.sum(Booking.total_price).label('total_revenue')
                )
                .group_by(extract('year', Booking.created_at))
                .order_by(extract('year', Booking.created_at).desc())
                .all()
            )
            
            for row in results:
                stats.append({
                    'period': str(int(row.year)),
                    'label': f"Năm {int(row.year)}",
                    'bookings_count': int(row.count),
                    'confirmed_revenue': float(row.confirmed_revenue or 0),
                    'total_revenue': float(row.total_revenue or 0)
                })

        return jsonify({
            'success': True,
            'period': period,
            'stats': stats
        }), 200

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error computing revenue stats: {str(e)}'}), 500


@seller_bp.route('/export/revenue', methods=['GET'])
@jwt_required()
def export_revenue():
    """Export revenue report to Excel or CSV"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        period = request.args.get('period', 'month')  # day, month, year
        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        export_format = request.args.get('format', 'excel').lower()

        # Get bookings
        if user.role == 'admin':
            bookings_query = db.session.query(Booking)
        else:
            bookings_query = db.session.query(Booking).join(Tour, Booking.tour_id == Tour.id).filter(Tour.seller_id == current_user_id)

        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                bookings_query = bookings_query.filter(Booking.created_at >= start_dt)
            except ValueError:
                pass

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                end_dt = end_dt + timedelta(days=1)
                bookings_query = bookings_query.filter(Booking.created_at < end_dt)
            except ValueError:
                pass

        bookings = bookings_query.order_by(Booking.created_at.desc()).all()

        if export_format == 'excel' and EXCEL_AVAILABLE:
            return export_revenue_to_excel(bookings, period, start_date, end_date)
        else:
            return export_revenue_to_csv(bookings, period, start_date, end_date)

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error exporting revenue: {str(e)}'}), 500


@seller_bp.route('/export/bookings', methods=['GET'])
@jwt_required()
def export_bookings():
    """Export bookings list to Excel or CSV"""
    try:
        current_user_id = get_jwt_identity()
        user = User.query.get(current_user_id)
        if not user:
            return jsonify({'error': 'User not found'}), 404

        start_date = request.args.get('start_date')
        end_date = request.args.get('end_date')
        export_format = request.args.get('format', 'excel').lower()

        # Get bookings for seller's tours
        if user.role == 'admin':
            bookings_query = db.session.query(Booking).join(Tour)
        else:
            bookings_query = db.session.query(Booking).join(Tour).filter(Tour.seller_id == current_user_id)

        if start_date:
            try:
                start_dt = datetime.strptime(start_date, '%Y-%m-%d')
                bookings_query = bookings_query.filter(Booking.created_at >= start_dt)
            except ValueError:
                pass

        if end_date:
            try:
                end_dt = datetime.strptime(end_date, '%Y-%m-%d')
                end_dt = end_dt + timedelta(days=1)
                bookings_query = bookings_query.filter(Booking.created_at < end_dt)
            except ValueError:
                pass

        bookings = bookings_query.order_by(Booking.created_at.desc()).all()

        if export_format == 'excel' and EXCEL_AVAILABLE:
            return export_bookings_to_excel(bookings, start_date, end_date)
        else:
            return export_bookings_to_csv(bookings, start_date, end_date)

    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error exporting bookings: {str(e)}'}), 500


def export_revenue_to_excel(bookings, period, start_date, end_date):
    """Export revenue to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Báo cáo doanh thu"
    
    # Set column widths
    ws.column_dimensions['A'].width = 15
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 20
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws['A1'] = "BÁO CÁO DOANH THU"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:H1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    row = 2
    if start_date or end_date:
        date_range = f"Từ {start_date or 'đầu'} đến {end_date or 'cuối'}"
        ws[f'A{row}'] = date_range
        ws.merge_cells(f'A{row}:H{row}')
        row += 1
    
    # Headers
    headers = ['Ngày đặt', 'Tour', 'Khách hàng', 'Số người', 'Tổng tiền', 'Trạng thái', 'Thanh toán', 'Ghi chú']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    row += 1
    
    # Add booking data
    total_revenue = 0
    for booking in bookings:
        ws.cell(row=row, column=1).value = booking.created_at.strftime('%d/%m/%Y %H:%M') if booking.created_at else ''
        ws.cell(row=row, column=2).value = booking.tour.title if booking.tour else 'N/A'
        ws.cell(row=row, column=3).value = booking.user.full_name if booking.user else booking.full_name or 'N/A'
        ws.cell(row=row, column=4).value = booking.participants
        ws.cell(row=row, column=5).value = booking.total_price
        ws.cell(row=row, column=6).value = booking.status
        ws.cell(row=row, column=7).value = booking.payment_status
        ws.cell(row=row, column=8).value = booking.notes or ''
        total_revenue += booking.total_price
        row += 1
    
    # Add summary
    row += 1
    ws.cell(row=row, column=4).value = "TỔNG DOANH THU:"
    ws.cell(row=row, column=4).font = Font(bold=True)
    ws.cell(row=row, column=5).value = total_revenue
    ws.cell(row=row, column=5).font = Font(bold=True)
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"revenue_report_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def export_revenue_to_csv(bookings, period, start_date, end_date):
    """Export revenue to CSV file"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write title
    writer.writerow(['BÁO CÁO DOANH THU'])
    if start_date or end_date:
        date_range = f"Từ {start_date or 'đầu'} đến {end_date or 'cuối'}"
        writer.writerow([date_range])
    writer.writerow([])
    
    # Column headers
    writer.writerow(['Ngày đặt', 'Tour', 'Khách hàng', 'Số người', 'Tổng tiền', 'Trạng thái', 'Thanh toán', 'Ghi chú'])
    
    # Booking data
    total_revenue = 0
    for booking in bookings:
        writer.writerow([
            booking.created_at.strftime('%d/%m/%Y %H:%M') if booking.created_at else '',
            booking.tour.title if booking.tour else 'N/A',
            booking.user.full_name if booking.user else booking.full_name or 'N/A',
            booking.participants,
            booking.total_price,
            booking.status,
            booking.payment_status,
            booking.notes or ''
        ])
        total_revenue += booking.total_price
    
    writer.writerow([])
    writer.writerow(['TỔNG DOANH THU:', '', '', '', total_revenue])
    
    # Create response
    output.seek(0)
    bytes_output = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    
    filename = f"revenue_report_{period}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        bytes_output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )


def export_bookings_to_excel(bookings, start_date, end_date):
    """Export bookings to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách đặt tour"
    
    # Set column widths
    ws.column_dimensions['A'].width = 10
    ws.column_dimensions['B'].width = 40
    ws.column_dimensions['C'].width = 25
    ws.column_dimensions['D'].width = 20
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 20
    ws.column_dimensions['G'].width = 15
    ws.column_dimensions['H'].width = 20
    ws.column_dimensions['I'].width = 15
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws['A1'] = "DANH SÁCH ĐẶT TOUR"
    ws['A1'].font = Font(bold=True, size=16)
    ws.merge_cells('A1:I1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    row = 2
    if start_date or end_date:
        date_range = f"Từ {start_date or 'đầu'} đến {end_date or 'cuối'}"
        ws[f'A{row}'] = date_range
        ws.merge_cells(f'A{row}:I{row}')
        row += 1
    
    # Headers
    headers = ['ID', 'Tour', 'Khách hàng', 'Email', 'SĐT', 'Ngày khởi hành', 'Số người', 'Tổng tiền', 'Trạng thái']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=row, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    row += 1
    
    # Add booking data
    for booking in bookings:
        ws.cell(row=row, column=1).value = booking.id
        ws.cell(row=row, column=2).value = booking.tour.title if booking.tour else 'N/A'
        ws.cell(row=row, column=3).value = booking.user.full_name if booking.user else booking.full_name or 'N/A'
        ws.cell(row=row, column=4).value = booking.email or (booking.user.email if booking.user else 'N/A')
        ws.cell(row=row, column=5).value = booking.phone or 'N/A'
        # booking.date is a string, not datetime
        ws.cell(row=row, column=6).value = booking.date if booking.date else 'N/A'
        ws.cell(row=row, column=7).value = booking.participants or 0
        ws.cell(row=row, column=8).value = booking.total_price or 0
        ws.cell(row=row, column=9).value = 'Đã xác nhận' if booking.status == 'confirmed' else 'Chờ xác nhận' if booking.status == 'pending' else 'Đã hủy'
        row += 1
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"bookings_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def export_bookings_to_csv(bookings, start_date, end_date):
    """Export bookings to CSV file"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write title
    writer.writerow(['DANH SÁCH ĐẶT TOUR'])
    if start_date or end_date:
        date_range = f"Từ {start_date or 'đầu'} đến {end_date or 'cuối'}"
        writer.writerow([date_range])
    writer.writerow([])
    
    # Column headers
    writer.writerow(['ID', 'Tour', 'Khách hàng', 'Email', 'SĐT', 'Ngày khởi hành', 'Số người', 'Tổng tiền', 'Trạng thái'])
    
    # Booking data
    for booking in bookings:
        writer.writerow([
            booking.id,
            booking.tour.title if booking.tour else 'N/A',
            booking.user.full_name if booking.user else booking.full_name or 'N/A',
            booking.email or (booking.user.email if booking.user else 'N/A'),
            booking.phone or 'N/A',
            booking.date if booking.date else 'N/A',  # booking.date is a string, not datetime
            booking.participants or 0,
            booking.total_price or 0,
            'Đã xác nhận' if booking.status == 'confirmed' else 'Chờ xác nhận' if booking.status == 'pending' else 'Đã hủy'
        ])
    
    # Create response
    output.seek(0)
    bytes_output = io.BytesIO(output.getvalue().encode('utf-8-sig'))
    
    filename = f"bookings_list_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    
    return send_file(
        bytes_output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )