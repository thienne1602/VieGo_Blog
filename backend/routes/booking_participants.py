from flask import Blueprint, request, jsonify, send_file
from flask_jwt_extended import jwt_required, get_jwt_identity
from models import db
from models.user import User
from models.booking import Booking
from models.tour import Tour
from models.booking_participant import BookingParticipant
from datetime import datetime
import io
import csv

# Try to import openpyxl for Excel export
try:
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("⚠️  openpyxl not available. Excel export will be disabled.")

booking_participants_bp = Blueprint('booking_participants', __name__, url_prefix='/api/booking-participants')


@booking_participants_bp.route('', methods=['GET'])
@jwt_required()
def get_participants_by_query():
    """Get all participants for a booking using query parameter"""
    try:
        booking_id = request.args.get('booking_id')
        if not booking_id:
            return jsonify({'error': 'booking_id query parameter is required'}), 400
        
        try:
            booking_id = int(booking_id)
        except ValueError:
            return jsonify({'error': 'booking_id must be a valid integer'}), 400
        
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        booking = Booking.query.get(booking_id)
        if not booking:
            return jsonify({'error': 'Booking not found'}), 404
        
        # Permission check: customer, seller, tour guide, or admin
        tour = Tour.query.get(booking.tour_id)
        from models.tour_assignment import TourAssignment
        assignment = TourAssignment.query.filter_by(booking_id=booking_id).first()
        
        is_customer = booking.user_id == current_user_id
        is_seller = tour and tour.seller_id == current_user_id
        is_guide = assignment and assignment.tour_guide_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_customer or is_seller or is_guide or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        participants = BookingParticipant.query.filter_by(booking_id=booking_id).all()
        
        return jsonify({
            'participants': [p.to_dict() for p in participants],
            'total': len(participants)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error fetching participants: {str(e)}'}), 500


@booking_participants_bp.route('/booking/<int:booking_id>', methods=['GET'])
@jwt_required()
def get_participants(booking_id):
    """Get all participants for a booking"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        booking = Booking.query.get(booking_id)
        if not booking:
            return jsonify({'error': 'Booking not found'}), 404
        
        # Permission check: customer, seller, tour guide, or admin
        tour = Tour.query.get(booking.tour_id)
        from models.tour_assignment import TourAssignment
        assignment = TourAssignment.query.filter_by(booking_id=booking_id).first()
        
        is_customer = booking.user_id == current_user_id
        is_seller = tour and tour.seller_id == current_user_id
        is_guide = assignment and assignment.tour_guide_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_customer or is_seller or is_guide or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        participants = BookingParticipant.query.filter_by(booking_id=booking_id).all()
        
        return jsonify({
            'participants': [p.to_dict() for p in participants],
            'total': len(participants)
        }), 200
        
    except Exception as e:
        return jsonify({'error': f'Error fetching participants: {str(e)}'}), 500


@booking_participants_bp.route('', methods=['POST'])
@jwt_required()
def create_participant():
    """Add a participant to a booking"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        data = request.get_json()
        
        if not data or 'booking_id' not in data or 'full_name' not in data:
            return jsonify({'error': 'booking_id and full_name are required'}), 400
        
        booking_id = data['booking_id']
        booking = Booking.query.get(booking_id)
        
        if not booking:
            return jsonify({'error': 'Booking not found'}), 404
        
        # Permission check: customer who made the booking, seller, or admin
        tour = Tour.query.get(booking.tour_id)
        
        is_customer = booking.user_id == current_user_id
        is_seller = tour and tour.seller_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_customer or is_seller or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        # Create participant
        participant = BookingParticipant(
            booking_id=booking_id,
            full_name=data['full_name'],
            gender=data.get('gender'),
            id_number=data.get('id_number'),
            passport_number=data.get('passport_number'),
            phone=data.get('phone'),
            email=data.get('email'),
            address=data.get('address'),
            participant_type=data.get('participant_type', 'adult'),
            special_requirements=data.get('special_requirements'),
            emergency_contact_name=data.get('emergency_contact_name'),
            emergency_contact_phone=data.get('emergency_contact_phone'),
            emergency_contact_relationship=data.get('emergency_contact_relationship')
        )
        
        # Parse date of birth if provided
        if 'date_of_birth' in data and data['date_of_birth']:
            try:
                participant.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except:
                pass
        
        db.session.add(participant)
        db.session.commit()
        
        return jsonify({
            'message': 'Participant added successfully',
            'participant': participant.to_dict()
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error creating participant: {str(e)}'}), 500


@booking_participants_bp.route('/booking/<int:booking_id>/batch', methods=['POST'])
@jwt_required()
def create_participants_batch(booking_id):
    """Add multiple participants to a booking at once"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        booking = Booking.query.get(booking_id)
        if not booking:
            return jsonify({'error': 'Booking not found'}), 404
        
        # Permission check
        tour = Tour.query.get(booking.tour_id)
        is_customer = booking.user_id == current_user_id
        is_seller = tour and tour.seller_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_customer or is_seller or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        if not data or 'participants' not in data or not isinstance(data['participants'], list):
            return jsonify({'error': 'participants array is required'}), 400
        
        # Clear existing participants if replace flag is set
        if data.get('replace', False):
            BookingParticipant.query.filter_by(booking_id=booking_id).delete()
        
        created_participants = []
        
        for p_data in data['participants']:
            if 'full_name' not in p_data:
                continue
            
            participant = BookingParticipant(
                booking_id=booking_id,
                full_name=p_data['full_name'],
                gender=p_data.get('gender'),
                id_number=p_data.get('id_number'),
                passport_number=p_data.get('passport_number'),
                phone=p_data.get('phone'),
                email=p_data.get('email'),
                address=p_data.get('address'),
                participant_type=p_data.get('participant_type', 'adult'),
                special_requirements=p_data.get('special_requirements'),
                emergency_contact_name=p_data.get('emergency_contact_name'),
                emergency_contact_phone=p_data.get('emergency_contact_phone'),
                emergency_contact_relationship=p_data.get('emergency_contact_relationship')
            )
            
            if 'date_of_birth' in p_data and p_data['date_of_birth']:
                try:
                    participant.date_of_birth = datetime.strptime(p_data['date_of_birth'], '%Y-%m-%d').date()
                except:
                    pass
            
            db.session.add(participant)
            created_participants.append(participant)
        
        db.session.commit()
        
        return jsonify({
            'message': f'{len(created_participants)} participants added successfully',
            'participants': [p.to_dict() for p in created_participants]
        }), 201
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error creating participants: {str(e)}'}), 500


@booking_participants_bp.route('/<int:participant_id>', methods=['PATCH'])
@jwt_required()
def update_participant(participant_id):
    """Update participant information"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        participant = BookingParticipant.query.get(participant_id)
        if not participant:
            return jsonify({'error': 'Participant not found'}), 404
        
        booking = Booking.query.get(participant.booking_id)
        tour = Tour.query.get(booking.tour_id) if booking else None
        
        # Permission check
        is_customer = booking and booking.user_id == current_user_id
        is_seller = tour and tour.seller_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_customer or is_seller or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        data = request.get_json()
        
        # Update fields
        if 'full_name' in data:
            participant.full_name = data['full_name']
        if 'gender' in data:
            participant.gender = data['gender']
        if 'date_of_birth' in data and data['date_of_birth']:
            try:
                participant.date_of_birth = datetime.strptime(data['date_of_birth'], '%Y-%m-%d').date()
            except:
                pass
        if 'id_number' in data:
            participant.id_number = data['id_number']
        if 'passport_number' in data:
            participant.passport_number = data['passport_number']
        if 'phone' in data:
            participant.phone = data['phone']
        if 'email' in data:
            participant.email = data['email']
        if 'address' in data:
            participant.address = data['address']
        if 'participant_type' in data:
            participant.participant_type = data['participant_type']
        if 'special_requirements' in data:
            participant.special_requirements = data['special_requirements']
        if 'emergency_contact_name' in data:
            participant.emergency_contact_name = data['emergency_contact_name']
        if 'emergency_contact_phone' in data:
            participant.emergency_contact_phone = data['emergency_contact_phone']
        if 'emergency_contact_relationship' in data:
            participant.emergency_contact_relationship = data['emergency_contact_relationship']
        
        participant.updated_at = datetime.utcnow()
        db.session.commit()
        
        return jsonify({
            'message': 'Participant updated successfully',
            'participant': participant.to_dict()
        }), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error updating participant: {str(e)}'}), 500


@booking_participants_bp.route('/<int:participant_id>', methods=['DELETE'])
@jwt_required()
def delete_participant(participant_id):
    """Delete a participant"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        participant = BookingParticipant.query.get(participant_id)
        if not participant:
            return jsonify({'error': 'Participant not found'}), 404
        
        booking = Booking.query.get(participant.booking_id)
        tour = Tour.query.get(booking.tour_id) if booking else None
        
        # Permission check
        is_customer = booking and booking.user_id == current_user_id
        is_seller = tour and tour.seller_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_customer or is_seller or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        db.session.delete(participant)
        db.session.commit()
        
        return jsonify({'message': 'Participant deleted successfully'}), 200
        
    except Exception as e:
        db.session.rollback()
        return jsonify({'error': f'Error deleting participant: {str(e)}'}), 500


@booking_participants_bp.route('/booking/<int:booking_id>/export', methods=['GET'])
@jwt_required()
def export_participants(booking_id):
    """Export participants list to Excel or CSV"""
    try:
        current_user_id = int(get_jwt_identity())
        user = User.query.get(current_user_id)
        
        if not user:
            return jsonify({'error': 'User not found'}), 404
        
        booking = Booking.query.get(booking_id)
        if not booking:
            return jsonify({'error': 'Booking not found'}), 404
        
        # Permission check: seller, tour guide, or admin
        tour = Tour.query.get(booking.tour_id)
        from models.tour_assignment import TourAssignment
        assignment = TourAssignment.query.filter_by(booking_id=booking_id).first()
        
        is_seller = tour and tour.seller_id == current_user_id
        is_guide = assignment and assignment.tour_guide_id == current_user_id
        is_admin = user.role == 'admin'
        
        if not (is_seller or is_guide or is_admin):
            return jsonify({'error': 'Unauthorized'}), 403
        
        participants = BookingParticipant.query.filter_by(booking_id=booking_id).all()
        
        # Get export format from query parameter (default: excel)
        export_format = request.args.get('format', 'excel').lower()
        
        if export_format == 'excel' and EXCEL_AVAILABLE:
            return export_to_excel(participants, booking, tour)
        else:
            return export_to_csv(participants, booking, tour)
        
    except Exception as e:
        return jsonify({'error': f'Error exporting participants: {str(e)}'}), 500


def export_to_excel(participants, booking, tour):
    """Export participants to Excel file"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Danh sách người tham gia"
    
    # Set column widths
    ws.column_dimensions['A'].width = 5
    ws.column_dimensions['B'].width = 25
    ws.column_dimensions['C'].width = 10
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 15
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 25
    ws.column_dimensions['H'].width = 30
    ws.column_dimensions['I'].width = 12
    ws.column_dimensions['J'].width = 30
    
    # Header styling
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Add title
    ws['A1'] = f"DANH SÁCH NGƯỜI THAM GIA TOUR"
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:J1')
    ws['A1'].alignment = Alignment(horizontal="center")
    
    # Add tour info
    ws['A2'] = f"Tour: {tour.title}"
    ws.merge_cells('A2:J2')
    ws['A3'] = f"Ngày khởi hành: {booking.date}"
    ws.merge_cells('A3:J3')
    ws['A4'] = f"Booking ID: {booking.id}"
    ws.merge_cells('A4:J4')
    
    # Headers
    headers = ['STT', 'Họ và tên', 'Giới tính', 'Ngày sinh', 'CMND/CCCD', 'Hộ chiếu', 
               'Điện thoại', 'Email', 'Loại', 'Yêu cầu đặc biệt']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=6, column=col)
        cell.value = header
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
    
    # Add participant data
    for idx, participant in enumerate(participants, 1):
        row = idx + 6
        ws.cell(row=row, column=1).value = idx
        ws.cell(row=row, column=2).value = participant.full_name
        ws.cell(row=row, column=3).value = participant.gender or ''
        ws.cell(row=row, column=4).value = participant.date_of_birth.strftime('%d/%m/%Y') if participant.date_of_birth else ''
        ws.cell(row=row, column=5).value = participant.id_number or ''
        ws.cell(row=row, column=6).value = participant.passport_number or ''
        ws.cell(row=row, column=7).value = participant.phone or ''
        ws.cell(row=row, column=8).value = participant.email or ''
        ws.cell(row=row, column=9).value = participant.participant_type or ''
        ws.cell(row=row, column=10).value = participant.special_requirements or ''
    
    # Save to BytesIO
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"participants_booking_{booking.id}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


def export_to_csv(participants, booking, tour):
    """Export participants to CSV file"""
    output = io.StringIO()
    writer = csv.writer(output)
    
    # Write headers
    writer.writerow(['DANH SÁCH NGƯỜI THAM GIA TOUR'])
    writer.writerow([f'Tour: {tour.title}'])
    writer.writerow([f'Ngày khởi hành: {booking.date}'])
    writer.writerow([f'Booking ID: {booking.id}'])
    writer.writerow([])
    
    # Column headers
    writer.writerow(['STT', 'Họ và tên', 'Giới tính', 'Ngày sinh', 'CMND/CCCD', 'Hộ chiếu',
                    'Điện thoại', 'Email', 'Loại', 'Yêu cầu đặc biệt'])
    
    # Participant data
    for idx, participant in enumerate(participants, 1):
        writer.writerow([
            idx,
            participant.full_name,
            participant.gender or '',
            participant.date_of_birth.strftime('%d/%m/%Y') if participant.date_of_birth else '',
            participant.id_number or '',
            participant.passport_number or '',
            participant.phone or '',
            participant.email or '',
            participant.participant_type or '',
            participant.special_requirements or ''
        ])
    
    # Create response
    output.seek(0)
    bytes_output = io.BytesIO(output.getvalue().encode('utf-8-sig'))  # UTF-8 with BOM for Excel compatibility
    
    filename = f"participants_booking_{booking.id}.csv"
    
    return send_file(
        bytes_output,
        mimetype='text/csv',
        as_attachment=True,
        download_name=filename
    )
