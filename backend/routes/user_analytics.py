"""
User Analytics Routes for Admin Dashboard
Thống kê dữ liệu người dùng, biểu đồ xu hướng, phân tích AI
"""

from flask import Blueprint, request, jsonify
from flask_jwt_extended import jwt_required, get_jwt_identity
from functools import wraps
from datetime import datetime, timedelta
from sqlalchemy import func, desc, and_, or_, case, extract
from collections import defaultdict
import json

from models import db, User, UserBehavior, UserInterestProfile, Post, Comment, Booking, Tour, UserPreferences

user_analytics_bp = Blueprint('user_analytics', __name__, url_prefix='/api/admin/user-analytics')

# Admin authorization decorator
def admin_required(f):
    @wraps(f)
    @jwt_required()
    def decorated_function(*args, **kwargs):
        try:
            current_user_id = get_jwt_identity()
            user_id_int = int(current_user_id) if isinstance(current_user_id, str) else current_user_id
            user = User.query.get(user_id_int)
            
            if not user:
                return jsonify({'error': 'User not found'}), 404
            
            if user.role not in ['admin', 'moderator']:
                return jsonify({'error': 'Admin access required'}), 403
            
            return f(*args, **kwargs)
        except Exception as e:
            return jsonify({'error': f'Authorization error: {str(e)}'}), 500
    return decorated_function


# ============================================================================
# THỐNG KÊ DỮ LIỆU NGƯỜI DÙNG
# ============================================================================

@user_analytics_bp.route('/overview', methods=['GET'])
@admin_required
def get_user_analytics_overview():
    """Tổng quan thống kê người dùng"""
    try:
        days = request.args.get('days', 30, type=int)
        cutoff_date = datetime.utcnow() - timedelta(days=days)
        
        # Tổng số người dùng
        total_users = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()
        
        # Người dùng mới trong khoảng thời gian
        new_users = User.query.filter(User.created_at >= cutoff_date).count()
        
        # Phân loại người dùng theo role
        role_distribution = db.session.query(
            User.role,
            func.count(User.id).label('count')
        ).group_by(User.role).all()
        
        role_stats = {role: count for role, count in role_distribution}
        
        # Người dùng đã xác thực email
        verified_users = User.query.filter_by(email_verified=True).count()
        
        # Người dùng có booking
        users_with_bookings = db.session.query(func.count(func.distinct(Booking.user_id))).scalar()
        
        # Thống kê engagement
        # - Users với posts
        users_with_posts = db.session.query(func.count(func.distinct(Post.author_id))).scalar()
        # - Users với comments
        users_with_comments = db.session.query(func.count(func.distinct(Comment.author_id))).scalar()
        
        # Người dùng hoạt động trong 7 ngày qua
        week_ago = datetime.utcnow() - timedelta(days=7)
        weekly_active_users = db.session.query(func.count(func.distinct(UserBehavior.user_id))).filter(
            UserBehavior.created_at >= week_ago
        ).scalar() or 0
        
        # Người dùng hoạt động hàng ngày (24h)
        day_ago = datetime.utcnow() - timedelta(days=1)
        daily_active_users = db.session.query(func.count(func.distinct(UserBehavior.user_id))).filter(
            UserBehavior.created_at >= day_ago
        ).scalar() or 0
        
        # Tỷ lệ chuyển đổi (users có booking / total users)
        conversion_rate = (users_with_bookings / total_users * 100) if total_users > 0 else 0
        
        # Tỷ lệ giữ chân (users hoạt động / total users)
        retention_rate = (weekly_active_users / total_users * 100) if total_users > 0 else 0
        
        return jsonify({
            'overview': {
                'totalUsers': total_users,
                'activeUsers': active_users,
                'newUsers': new_users,
                'verifiedUsers': verified_users,
                'dailyActiveUsers': daily_active_users,
                'weeklyActiveUsers': weekly_active_users,
                'usersWithBookings': users_with_bookings,
                'usersWithPosts': users_with_posts,
                'usersWithComments': users_with_comments,
                'conversionRate': round(conversion_rate, 2),
                'retentionRate': round(retention_rate, 2)
            },
            'roleDistribution': role_stats,
            'period': {
                'days': days,
                'startDate': cutoff_date.isoformat(),
                'endDate': datetime.utcnow().isoformat()
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@user_analytics_bp.route('/trends', methods=['GET'])
@admin_required
def get_user_trends():
    """Xu hướng người dùng theo thời gian"""
    try:
        days = request.args.get('days', 30, type=int)
        granularity = request.args.get('granularity', 'day')  # day, week, month
        cutoff_date = datetime.utcnow() - timedelta(days=days)
        
        # Xu hướng đăng ký mới theo ngày
        if granularity == 'day':
            registration_trend = db.session.query(
                func.date(User.created_at).label('date'),
                func.count(User.id).label('count')
            ).filter(
                User.created_at >= cutoff_date
            ).group_by(
                func.date(User.created_at)
            ).order_by(
                func.date(User.created_at)
            ).all()
        elif granularity == 'week':
            registration_trend = db.session.query(
                func.yearweek(User.created_at).label('week'),
                func.count(User.id).label('count')
            ).filter(
                User.created_at >= cutoff_date
            ).group_by(
                func.yearweek(User.created_at)
            ).order_by(
                func.yearweek(User.created_at)
            ).all()
        else:  # month
            registration_trend = db.session.query(
                func.date_format(User.created_at, '%Y-%m').label('month'),
                func.count(User.id).label('count')
            ).filter(
                User.created_at >= cutoff_date
            ).group_by(
                func.date_format(User.created_at, '%Y-%m')
            ).order_by(
                func.date_format(User.created_at, '%Y-%m')
            ).all()
        
        # Xu hướng hoạt động người dùng
        if granularity == 'day':
            activity_trend = db.session.query(
                func.date(UserBehavior.created_at).label('date'),
                func.count(func.distinct(UserBehavior.user_id)).label('active_users'),
                func.count(UserBehavior.id).label('total_actions')
            ).filter(
                UserBehavior.created_at >= cutoff_date
            ).group_by(
                func.date(UserBehavior.created_at)
            ).order_by(
                func.date(UserBehavior.created_at)
            ).all()
        else:
            activity_trend = db.session.query(
                func.date_format(UserBehavior.created_at, '%Y-%m').label('month'),
                func.count(func.distinct(UserBehavior.user_id)).label('active_users'),
                func.count(UserBehavior.id).label('total_actions')
            ).filter(
                UserBehavior.created_at >= cutoff_date
            ).group_by(
                func.date_format(UserBehavior.created_at, '%Y-%m')
            ).order_by(
                func.date_format(UserBehavior.created_at, '%Y-%m')
            ).all()
        
        # Xu hướng booking
        booking_trend = db.session.query(
            func.date(Booking.created_at).label('date'),
            func.count(Booking.id).label('bookings'),
            func.sum(Booking.total_price).label('revenue')
        ).filter(
            Booking.created_at >= cutoff_date
        ).group_by(
            func.date(Booking.created_at)
        ).order_by(
            func.date(Booking.created_at)
        ).all()
        
        return jsonify({
            'registrationTrend': [
                {'date': str(r.date if hasattr(r, 'date') else r[0]), 'count': r.count}
                for r in registration_trend
            ],
            'activityTrend': [
                {
                    'date': str(a.date if hasattr(a, 'date') else a[0]),
                    'activeUsers': a.active_users,
                    'totalActions': a.total_actions
                }
                for a in activity_trend
            ],
            'bookingTrend': [
                {
                    'date': str(b.date),
                    'bookings': b.bookings,
                    'revenue': float(b.revenue or 0)
                }
                for b in booking_trend
            ],
            'period': {
                'days': days,
                'granularity': granularity
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@user_analytics_bp.route('/demographics', methods=['GET'])
@admin_required
def get_user_demographics():
    """Phân tích nhân khẩu học người dùng"""
    try:
        # Phân bố theo vị trí
        location_distribution = db.session.query(
            User.location,
            func.count(User.id).label('count')
        ).filter(
            User.location.isnot(None),
            User.location != ''
        ).group_by(User.location).order_by(desc('count')).limit(20).all()
        
        # Phân bố theo ngôn ngữ
        language_distribution = db.session.query(
            User.language,
            func.count(User.id).label('count')
        ).group_by(User.language).all()
        
        # Phân bố theo level
        level_distribution = db.session.query(
            User.level,
            func.count(User.id).label('count')
        ).group_by(User.level).order_by(User.level).all()
        
        # Phân bố thiết bị (từ behaviors)
        device_distribution = db.session.query(
            UserBehavior.device_type,
            func.count(func.distinct(UserBehavior.user_id)).label('count')
        ).filter(
            UserBehavior.device_type.isnot(None)
        ).group_by(UserBehavior.device_type).all()
        
        return jsonify({
            'locationDistribution': [
                {'location': loc or 'Không xác định', 'count': count}
                for loc, count in location_distribution
            ],
            'languageDistribution': [
                {'language': lang or 'vi', 'count': count}
                for lang, count in language_distribution
            ],
            'levelDistribution': [
                {'level': level, 'count': count}
                for level, count in level_distribution
            ],
            'deviceDistribution': [
                {'device': device or 'desktop', 'count': count}
                for device, count in device_distribution
            ]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# PHÂN TÍCH SỞ THÍCH (AI ANALYSIS)
# ============================================================================

@user_analytics_bp.route('/interests-analysis', methods=['GET'])
@admin_required
def get_interests_analysis():
    """Phân tích sở thích toàn bộ người dùng bằng AI"""
    try:
        # Lấy tất cả interest profiles
        profiles = UserInterestProfile.query.all()
        
        # Tính trung bình điểm cho mỗi danh mục
        category_totals = defaultdict(lambda: {'sum': 0, 'count': 0})
        engagement_counts = defaultdict(int)
        price_data = []
        duration_data = []
        
        for profile in profiles:
            # Điểm danh mục
            category_totals['adventure']['sum'] += profile.adventure_score or 0
            category_totals['adventure']['count'] += 1
            
            category_totals['cultural']['sum'] += profile.cultural_score or 0
            category_totals['cultural']['count'] += 1
            
            category_totals['food']['sum'] += profile.food_score or 0
            category_totals['food']['count'] += 1
            
            category_totals['nature']['sum'] += profile.nature_score or 0
            category_totals['nature']['count'] += 1
            
            category_totals['urban']['sum'] += profile.urban_score or 0
            category_totals['urban']['count'] += 1
            
            category_totals['spiritual']['sum'] += profile.spiritual_score or 0
            category_totals['spiritual']['count'] += 1
            
            # Engagement level
            engagement_counts[profile.engagement_level] += 1
            
            # Price và duration
            price_data.append({
                'min': profile.preferred_price_min,
                'max': profile.preferred_price_max,
                'sensitivity': profile.price_sensitivity
            })
            
            duration_data.append({
                'min': profile.preferred_duration_min,
                'max': profile.preferred_duration_max
            })
        
        # Tính trung bình
        category_averages = {}
        for cat, data in category_totals.items():
            category_averages[cat] = round(data['sum'] / data['count'], 2) if data['count'] > 0 else 0
        
        # Tính phân phối giá
        avg_price_min = sum(p['min'] for p in price_data) / len(price_data) if price_data else 0
        avg_price_max = sum(p['max'] for p in price_data) / len(price_data) if price_data else 0
        avg_sensitivity = sum(p['sensitivity'] for p in price_data) / len(price_data) if price_data else 0.5
        
        # Tính phân phối thời gian
        avg_duration_min = sum(d['min'] for d in duration_data) / len(duration_data) if duration_data else 1
        avg_duration_max = sum(d['max'] for d in duration_data) / len(duration_data) if duration_data else 7
        
        # Lấy top tags phổ biến
        tag_counts = defaultdict(int)
        for profile in profiles:
            for tag in profile.get_favorite_tags():
                tag_counts[tag] += 1
        
        top_tags = sorted(tag_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        
        # Lấy top địa điểm yêu thích
        location_counts = defaultdict(int)
        for profile in profiles:
            for loc in profile.get_preferred_locations():
                location_counts[loc] += 1
        
        top_locations = sorted(location_counts.items(), key=lambda x: x[1], reverse=True)[:20]
        
        return jsonify({
            'totalProfiles': len(profiles),
            'categoryInterests': category_averages,
            'engagementDistribution': dict(engagement_counts),
            'pricePreferences': {
                'averageMin': round(avg_price_min, 0),
                'averageMax': round(avg_price_max, 0),
                'averageSensitivity': round(avg_sensitivity, 2)
            },
            'durationPreferences': {
                'averageMin': round(avg_duration_min, 1),
                'averageMax': round(avg_duration_max, 1)
            },
            'topTags': [{'tag': tag, 'count': count} for tag, count in top_tags],
            'topLocations': [{'location': loc, 'count': count} for loc, count in top_locations]
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@user_analytics_bp.route('/analyze-all-users', methods=['POST'])
@admin_required
def analyze_all_users():
    """Phân tích AI cho tất cả người dùng (batch process)"""
    try:
        from utils.ai_analytics import get_analytics_service, init_analytics_service
        
        analytics = get_analytics_service()
        if not analytics:
            analytics = init_analytics_service(db)
        
        days = request.json.get('days', 365) if request.json else 365
        
        # Lấy tất cả users active
        users = User.query.filter_by(is_active=True).all()
        
        results = {
            'total': len(users),
            'analyzed': 0,
            'failed': 0,
            'errors': []
        }
        
        for user in users:
            try:
                analytics.analyze_user_interests(user.id, days=days)
                results['analyzed'] += 1
            except Exception as e:
                results['failed'] += 1
                results['errors'].append({
                    'userId': user.id,
                    'error': str(e)
                })
        
        return jsonify({
            'message': 'User analysis completed',
            'results': results
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# PHÂN NHÓM NGƯỜI DÙNG (USER SEGMENTS)
# ============================================================================

@user_analytics_bp.route('/segments', methods=['GET'])
@admin_required
def get_user_segments():
    """Lấy các phân khúc người dùng đã định nghĩa"""
    try:
        segments = [
            {
                'id': 'high_engagement',
                'name': 'Người dùng tích cực',
                'description': 'Người dùng có engagement level cao hoặc rất cao',
                'criteria': {'engagement_level': ['high', 'very_high']}
            },
            {
                'id': 'adventure_lovers',
                'name': 'Yêu thích phiêu lưu',
                'description': 'Người dùng có điểm adventure_score > 50',
                'criteria': {'adventure_score_min': 50}
            },
            {
                'id': 'cultural_enthusiasts',
                'name': 'Yêu thích văn hóa',
                'description': 'Người dùng có điểm cultural_score > 50',
                'criteria': {'cultural_score_min': 50}
            },
            {
                'id': 'food_lovers',
                'name': 'Yêu thích ẩm thực',
                'description': 'Người dùng có điểm food_score > 50',
                'criteria': {'food_score_min': 50}
            },
            {
                'id': 'nature_explorers',
                'name': 'Yêu thiên nhiên',
                'description': 'Người dùng có điểm nature_score > 50',
                'criteria': {'nature_score_min': 50}
            },
            {
                'id': 'budget_travelers',
                'name': 'Du lịch tiết kiệm',
                'description': 'Người dùng nhạy cảm về giá (price_sensitivity > 0.7)',
                'criteria': {'price_sensitivity_min': 0.7}
            },
            {
                'id': 'premium_travelers',
                'name': 'Du lịch cao cấp',
                'description': 'Người dùng ít nhạy cảm về giá (price_sensitivity < 0.3)',
                'criteria': {'price_sensitivity_max': 0.3}
            },
            {
                'id': 'new_users',
                'name': 'Người dùng mới',
                'description': 'Đăng ký trong 30 ngày gần nhất',
                'criteria': {'registered_within_days': 30}
            },
            {
                'id': 'inactive_users',
                'name': 'Người dùng không hoạt động',
                'description': 'Không có hoạt động trong 30 ngày qua',
                'criteria': {'inactive_days': 30}
            },
            {
                'id': 'high_spenders',
                'name': 'Chi tiêu cao',
                'description': 'Tổng chi tiêu > 10.000.000 VND',
                'criteria': {'total_spent_min': 10000000}
            },
            {
                'id': 'frequent_bookers',
                'name': 'Đặt tour thường xuyên',
                'description': 'Có từ 3 booking trở lên',
                'criteria': {'total_bookings_min': 3}
            },
            {
                'id': 'email_engaged',
                'name': 'Tương tác email tốt',
                'description': 'Tỷ lệ mở email > 50%',
                'criteria': {'email_open_rate_min': 0.5}
            }
        ]
        
        # Đếm số user cho mỗi segment
        for segment in segments:
            segment['userCount'] = _count_segment_users(segment['criteria'])
        
        return jsonify({
            'segments': segments
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


def _count_segment_users(criteria):
    """Đếm số người dùng trong một segment"""
    try:
        query = db.session.query(User.id)
        
        # Join với UserInterestProfile nếu cần
        needs_profile = any(k in criteria for k in [
            'engagement_level', 'adventure_score_min', 'cultural_score_min',
            'food_score_min', 'nature_score_min', 'price_sensitivity_min',
            'price_sensitivity_max', 'total_spent_min', 'total_bookings_min',
            'email_open_rate_min'
        ])
        
        if needs_profile:
            query = query.outerjoin(UserInterestProfile, User.id == UserInterestProfile.user_id)
        
        # Áp dụng filters
        if 'engagement_level' in criteria:
            query = query.filter(UserInterestProfile.engagement_level.in_(criteria['engagement_level']))
        
        if 'adventure_score_min' in criteria:
            query = query.filter(UserInterestProfile.adventure_score >= criteria['adventure_score_min'])
        
        if 'cultural_score_min' in criteria:
            query = query.filter(UserInterestProfile.cultural_score >= criteria['cultural_score_min'])
        
        if 'food_score_min' in criteria:
            query = query.filter(UserInterestProfile.food_score >= criteria['food_score_min'])
        
        if 'nature_score_min' in criteria:
            query = query.filter(UserInterestProfile.nature_score >= criteria['nature_score_min'])
        
        if 'price_sensitivity_min' in criteria:
            query = query.filter(UserInterestProfile.price_sensitivity >= criteria['price_sensitivity_min'])
        
        if 'price_sensitivity_max' in criteria:
            query = query.filter(UserInterestProfile.price_sensitivity <= criteria['price_sensitivity_max'])
        
        if 'registered_within_days' in criteria:
            cutoff = datetime.utcnow() - timedelta(days=criteria['registered_within_days'])
            query = query.filter(User.created_at >= cutoff)
        
        if 'inactive_days' in criteria:
            cutoff = datetime.utcnow() - timedelta(days=criteria['inactive_days'])
            # Tìm users không có behavior nào trong khoảng thời gian
            active_user_ids = db.session.query(func.distinct(UserBehavior.user_id)).filter(
                UserBehavior.created_at >= cutoff
            ).scalar_subquery()
            query = query.filter(~User.id.in_(active_user_ids))
        
        if 'total_spent_min' in criteria:
            query = query.filter(UserInterestProfile.total_spent >= criteria['total_spent_min'])
        
        if 'total_bookings_min' in criteria:
            query = query.filter(UserInterestProfile.total_bookings >= criteria['total_bookings_min'])
        
        if 'email_open_rate_min' in criteria:
            query = query.filter(UserInterestProfile.email_open_rate >= criteria['email_open_rate_min'])
        
        return query.count()
    except Exception as e:
        return 0


@user_analytics_bp.route('/segments/<segment_id>/users', methods=['GET'])
@admin_required
def get_segment_users(segment_id):
    """Lấy danh sách người dùng trong một segment"""
    try:
        page = request.args.get('page', 1, type=int)
        per_page = request.args.get('per_page', 20, type=int)
        
        # Định nghĩa criteria cho segment
        segment_criteria = {
            'high_engagement': {'engagement_level': ['high', 'very_high']},
            'adventure_lovers': {'adventure_score_min': 50},
            'cultural_enthusiasts': {'cultural_score_min': 50},
            'food_lovers': {'food_score_min': 50},
            'nature_explorers': {'nature_score_min': 50},
            'budget_travelers': {'price_sensitivity_min': 0.7},
            'premium_travelers': {'price_sensitivity_max': 0.3},
            'new_users': {'registered_within_days': 30},
            'inactive_users': {'inactive_days': 30},
            'high_spenders': {'total_spent_min': 10000000},
            'frequent_bookers': {'total_bookings_min': 3},
            'email_engaged': {'email_open_rate_min': 0.5}
        }
        
        if segment_id not in segment_criteria:
            return jsonify({'error': 'Segment not found'}), 404
        
        criteria = segment_criteria[segment_id]
        
        # Build query
        query = User.query
        
        needs_profile = any(k in criteria for k in [
            'engagement_level', 'adventure_score_min', 'cultural_score_min',
            'food_score_min', 'nature_score_min', 'price_sensitivity_min',
            'price_sensitivity_max', 'total_spent_min', 'total_bookings_min',
            'email_open_rate_min'
        ])
        
        if needs_profile:
            query = query.outerjoin(UserInterestProfile, User.id == UserInterestProfile.user_id)
        
        # Áp dụng filters (tương tự hàm _count_segment_users)
        if 'engagement_level' in criteria:
            query = query.filter(UserInterestProfile.engagement_level.in_(criteria['engagement_level']))
        
        if 'adventure_score_min' in criteria:
            query = query.filter(UserInterestProfile.adventure_score >= criteria['adventure_score_min'])
        
        if 'cultural_score_min' in criteria:
            query = query.filter(UserInterestProfile.cultural_score >= criteria['cultural_score_min'])
        
        if 'food_score_min' in criteria:
            query = query.filter(UserInterestProfile.food_score >= criteria['food_score_min'])
        
        if 'nature_score_min' in criteria:
            query = query.filter(UserInterestProfile.nature_score >= criteria['nature_score_min'])
        
        if 'price_sensitivity_min' in criteria:
            query = query.filter(UserInterestProfile.price_sensitivity >= criteria['price_sensitivity_min'])
        
        if 'price_sensitivity_max' in criteria:
            query = query.filter(UserInterestProfile.price_sensitivity <= criteria['price_sensitivity_max'])
        
        if 'registered_within_days' in criteria:
            cutoff = datetime.utcnow() - timedelta(days=criteria['registered_within_days'])
            query = query.filter(User.created_at >= cutoff)
        
        if 'inactive_days' in criteria:
            cutoff = datetime.utcnow() - timedelta(days=criteria['inactive_days'])
            active_user_ids = db.session.query(func.distinct(UserBehavior.user_id)).filter(
                UserBehavior.created_at >= cutoff
            ).subquery()
            query = query.filter(~User.id.in_(active_user_ids))
        
        if 'total_spent_min' in criteria:
            query = query.filter(UserInterestProfile.total_spent >= criteria['total_spent_min'])
        
        if 'total_bookings_min' in criteria:
            query = query.filter(UserInterestProfile.total_bookings >= criteria['total_bookings_min'])
        
        if 'email_open_rate_min' in criteria:
            query = query.filter(UserInterestProfile.email_open_rate >= criteria['email_open_rate_min'])
        
        # Paginate
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        users = []
        for user in pagination.items:
            users.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'fullName': user.full_name,
                'avatarUrl': user.avatar_url,
                'role': user.role,
                'createdAt': user.created_at.isoformat()
            })
        
        return jsonify({
            'users': users,
            'pagination': {
                'currentPage': page,
                'perPage': per_page,
                'totalPages': pagination.pages,
                'totalItems': pagination.total
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# CUSTOM SEGMENT QUERY
# ============================================================================

@user_analytics_bp.route('/segments/query', methods=['POST'])
@admin_required
def query_custom_segment():
    """Query người dùng với criteria tùy chỉnh"""
    try:
        data = request.get_json()
        criteria = data.get('criteria', {})
        page = data.get('page', 1)
        per_page = data.get('per_page', 20)
        
        query = User.query.filter_by(is_active=True)
        
        # Filter by role
        if 'roles' in criteria and criteria['roles']:
            query = query.filter(User.role.in_(criteria['roles']))
        
        # Filter by location
        if 'locations' in criteria and criteria['locations']:
            query = query.filter(User.location.in_(criteria['locations']))
        
        # Filter by registration date
        if 'registered_after' in criteria:
            query = query.filter(User.created_at >= datetime.fromisoformat(criteria['registered_after']))
        if 'registered_before' in criteria:
            query = query.filter(User.created_at <= datetime.fromisoformat(criteria['registered_before']))
        
        # Filter by interests
        if any(k in criteria for k in ['min_adventure', 'min_cultural', 'min_food', 'min_nature',
                                        'min_engagement', 'min_bookings', 'min_spent']):
            query = query.outerjoin(UserInterestProfile, User.id == UserInterestProfile.user_id)
            
            if 'min_adventure' in criteria:
                query = query.filter(UserInterestProfile.adventure_score >= criteria['min_adventure'])
            if 'min_cultural' in criteria:
                query = query.filter(UserInterestProfile.cultural_score >= criteria['min_cultural'])
            if 'min_food' in criteria:
                query = query.filter(UserInterestProfile.food_score >= criteria['min_food'])
            if 'min_nature' in criteria:
                query = query.filter(UserInterestProfile.nature_score >= criteria['min_nature'])
            if 'min_bookings' in criteria:
                query = query.filter(UserInterestProfile.total_bookings >= criteria['min_bookings'])
            if 'min_spent' in criteria:
                query = query.filter(UserInterestProfile.total_spent >= criteria['min_spent'])
            if 'min_engagement' in criteria:
                engagement_levels = {
                    1: 'low',
                    2: 'medium',
                    3: 'high',
                    4: 'very_high'
                }
                min_level = criteria['min_engagement']
                allowed_levels = [v for k, v in engagement_levels.items() if k >= min_level]
                query = query.filter(UserInterestProfile.engagement_level.in_(allowed_levels))
        
        # Filter by email verified
        if 'email_verified' in criteria:
            query = query.filter(User.email_verified == criteria['email_verified'])
        
        # Paginate
        pagination = query.paginate(page=page, per_page=per_page, error_out=False)
        
        users = []
        for user in pagination.items:
            users.append({
                'id': user.id,
                'username': user.username,
                'email': user.email,
                'fullName': user.full_name,
                'avatarUrl': user.avatar_url,
                'role': user.role,
                'location': user.location,
                'emailVerified': user.email_verified,
                'createdAt': user.created_at.isoformat()
            })
        
        return jsonify({
            'users': users,
            'pagination': {
                'currentPage': page,
                'perPage': per_page,
                'totalPages': pagination.pages,
                'totalItems': pagination.total
            }
        }), 200
    except Exception as e:
        return jsonify({'error': str(e)}), 500


# ============================================================================
# XUẤT EXCEL
# ============================================================================

@user_analytics_bp.route('/export/users', methods=['GET'])
@admin_required
def export_users_excel():
    """Xuất danh sách người dùng ra Excel"""
    try:
        from io import BytesIO
        from flask import send_file
        
        # Optional: try to use openpyxl, fallback to CSV if not available
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            use_excel = True
        except ImportError:
            use_excel = False
        
        segment_id = request.args.get('segment')
        
        # Build query
        query = User.query.filter_by(is_active=True)
        
        if segment_id:
            # Apply segment filters
            query = query.outerjoin(UserInterestProfile, User.id == UserInterestProfile.user_id)
            # Add segment-specific filters based on segment_id
        
        users = query.order_by(User.created_at.desc()).all()
        
        if use_excel:
            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Người dùng"
            
            # Header style
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Headers
            headers = ['ID', 'Username', 'Email', 'Họ tên', 'Vai trò', 'Vị trí', 
                      'Điểm', 'Cấp độ', 'Email xác thực', 'Trạng thái', 'Ngày tạo']
            
            for col, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col, value=header)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            
            # Data rows
            for row, user in enumerate(users, 2):
                ws.cell(row=row, column=1, value=user.id).border = thin_border
                ws.cell(row=row, column=2, value=user.username).border = thin_border
                ws.cell(row=row, column=3, value=user.email).border = thin_border
                ws.cell(row=row, column=4, value=user.full_name or '').border = thin_border
                ws.cell(row=row, column=5, value=user.role).border = thin_border
                ws.cell(row=row, column=6, value=user.location or '').border = thin_border
                ws.cell(row=row, column=7, value=user.points or 0).border = thin_border
                ws.cell(row=row, column=8, value=user.level or 1).border = thin_border
                ws.cell(row=row, column=9, value='Có' if user.email_verified else 'Không').border = thin_border
                ws.cell(row=row, column=10, value='Hoạt động' if user.is_active else 'Khóa').border = thin_border
                ws.cell(row=row, column=11, value=user.created_at.strftime('%d/%m/%Y %H:%M') if user.created_at else '').border = thin_border
            
            # Auto-width columns (skip merged cells)
            from openpyxl.cell.cell import MergedCell
            from openpyxl.utils import get_column_letter
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
            
            # Save to bytes
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            return send_file(
                output,
                mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                as_attachment=True,
                download_name=filename
            )
        else:
            # Fallback to CSV
            import csv
            output = BytesIO()
            
            # Write BOM for Excel UTF-8 compatibility
            output.write(b'\xef\xbb\xbf')
            
            writer = csv.writer(output.getvalue().decode('utf-8').encode('utf-8'))
            
            # This is a simplified CSV export
            csv_data = "ID,Username,Email,Họ tên,Vai trò,Vị trí,Điểm,Cấp độ,Email xác thực,Trạng thái,Ngày tạo\n"
            for user in users:
                csv_data += f"{user.id},{user.username},{user.email},{user.full_name or ''},{user.role},{user.location or ''},{user.points or 0},{user.level or 1},{'Có' if user.email_verified else 'Không'},{'Hoạt động' if user.is_active else 'Khóa'},{user.created_at.strftime('%d/%m/%Y %H:%M') if user.created_at else ''}\n"
            
            output = BytesIO(csv_data.encode('utf-8-sig'))
            filename = f"users_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
            
            return send_file(
                output,
                mimetype='text/csv',
                as_attachment=True,
                download_name=filename
            )
            
    except Exception as e:
        return jsonify({'error': str(e)}), 500


@user_analytics_bp.route('/export/analytics', methods=['GET'])
@admin_required
def export_analytics_excel():
    """Xuất báo cáo phân tích người dùng ra Excel"""
    try:
        from io import BytesIO
        from flask import send_file
        
        try:
            from openpyxl import Workbook
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.chart import BarChart, PieChart, Reference
        except ImportError:
            return jsonify({'error': 'openpyxl library not installed. Run: pip install openpyxl'}), 500
        
        days = request.args.get('days', 30, type=int)
        cutoff_date = datetime.utcnow() - timedelta(days=days)
        
        wb = Workbook()
        
        # Style definitions
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        title_font = Font(bold=True, size=14)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ==================== Sheet 1: Tổng quan ====================
        ws1 = wb.active
        ws1.title = "Tổng quan"
        
        # Title
        ws1['A1'] = f"BÁO CÁO THỐNG KÊ NGƯỜI DÙNG - {days} NGÀY GẦN NHẤT"
        ws1['A1'].font = title_font
        ws1.merge_cells('A1:D1')
        
        ws1['A2'] = f"Ngày xuất: {datetime.now().strftime('%d/%m/%Y %H:%M')}"
        
        # Overview stats
        total_users = User.query.count()
        active_users = User.query.filter_by(is_active=True).count()
        new_users = User.query.filter(User.created_at >= cutoff_date).count()
        verified_users = User.query.filter_by(email_verified=True).count()
        
        overview_data = [
            ['Chỉ số', 'Giá trị'],
            ['Tổng người dùng', total_users],
            ['Người dùng hoạt động', active_users],
            ['Người dùng mới (trong kỳ)', new_users],
            ['Email đã xác thực', verified_users],
        ]
        
        for row_idx, row_data in enumerate(overview_data, 4):
            for col_idx, value in enumerate(row_data, 1):
                cell = ws1.cell(row=row_idx, column=col_idx, value=value)
                if row_idx == 4:
                    cell.font = header_font
                    cell.fill = header_fill
                cell.border = thin_border
        
        # ==================== Sheet 2: Phân bố vai trò ====================
        ws2 = wb.create_sheet("Phân bố vai trò")
        
        role_distribution = db.session.query(
            User.role,
            func.count(User.id).label('count')
        ).group_by(User.role).all()
        
        ws2['A1'] = "PHÂN BỐ NGƯỜI DÙNG THEO VAI TRÒ"
        ws2['A1'].font = title_font
        
        role_headers = ['Vai trò', 'Số lượng', 'Tỷ lệ (%)']
        for col, header in enumerate(role_headers, 1):
            cell = ws2.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        role_names = {
            'user': 'Người dùng',
            'seller': 'Người bán',
            'tour_guide': 'Hướng dẫn viên',
            'moderator': 'Kiểm duyệt viên',
            'admin': 'Quản trị viên'
        }
        
        for row_idx, (role, count) in enumerate(role_distribution, 4):
            ws2.cell(row=row_idx, column=1, value=role_names.get(role, role)).border = thin_border
            ws2.cell(row=row_idx, column=2, value=count).border = thin_border
            ws2.cell(row=row_idx, column=3, value=round(count/total_users*100, 2)).border = thin_border
        
        # ==================== Sheet 3: Xu hướng đăng ký ====================
        ws3 = wb.create_sheet("Xu hướng đăng ký")
        
        registration_trend = db.session.query(
            func.date(User.created_at).label('date'),
            func.count(User.id).label('count')
        ).filter(
            User.created_at >= cutoff_date
        ).group_by(
            func.date(User.created_at)
        ).order_by(
            func.date(User.created_at)
        ).all()
        
        ws3['A1'] = "XU HƯỚNG ĐĂNG KÝ NGƯỜI DÙNG MỚI"
        ws3['A1'].font = title_font
        
        trend_headers = ['Ngày', 'Số lượng đăng ký']
        for col, header in enumerate(trend_headers, 1):
            cell = ws3.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, (date, count) in enumerate(registration_trend, 4):
            ws3.cell(row=row_idx, column=1, value=str(date)).border = thin_border
            ws3.cell(row=row_idx, column=2, value=count).border = thin_border
        
        # ==================== Sheet 4: Phân tích sở thích ====================
        ws4 = wb.create_sheet("Phân tích sở thích")
        
        ws4['A1'] = "PHÂN TÍCH SỞ THÍCH NGƯỜI DÙNG (AI)"
        ws4['A1'].font = title_font
        
        # Get average interest scores
        interest_stats = db.session.query(
            func.avg(UserInterestProfile.adventure_score).label('adventure'),
            func.avg(UserInterestProfile.cultural_score).label('cultural'),
            func.avg(UserInterestProfile.food_score).label('food'),
            func.avg(UserInterestProfile.nature_score).label('nature'),
            func.avg(UserInterestProfile.urban_score).label('urban'),
            func.avg(UserInterestProfile.spiritual_score).label('spiritual')
        ).first()
        
        interest_headers = ['Danh mục', 'Điểm trung bình']
        for col, header in enumerate(interest_headers, 1):
            cell = ws4.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        categories = [
            ('Phiêu lưu', interest_stats.adventure or 0),
            ('Văn hóa', interest_stats.cultural or 0),
            ('Ẩm thực', interest_stats.food or 0),
            ('Thiên nhiên', interest_stats.nature or 0),
            ('Đô thị', interest_stats.urban or 0),
            ('Tâm linh', interest_stats.spiritual or 0)
        ]
        
        for row_idx, (name, score) in enumerate(categories, 4):
            ws4.cell(row=row_idx, column=1, value=name).border = thin_border
            ws4.cell(row=row_idx, column=2, value=round(score, 2)).border = thin_border
        
        # Engagement distribution
        ws4['A12'] = "PHÂN BỐ MỨC ĐỘ TƯƠNG TÁC"
        ws4['A12'].font = Font(bold=True, size=12)
        
        engagement_dist = db.session.query(
            UserInterestProfile.engagement_level,
            func.count(UserInterestProfile.id).label('count')
        ).group_by(UserInterestProfile.engagement_level).all()
        
        engagement_headers = ['Mức độ', 'Số lượng']
        for col, header in enumerate(engagement_headers, 1):
            cell = ws4.cell(row=14, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        engagement_names = {
            'very_high': 'Rất cao',
            'high': 'Cao',
            'medium': 'Trung bình',
            'low': 'Thấp'
        }
        
        for row_idx, (level, count) in enumerate(engagement_dist, 15):
            ws4.cell(row=row_idx, column=1, value=engagement_names.get(level, level)).border = thin_border
            ws4.cell(row=row_idx, column=2, value=count).border = thin_border
        
        # ==================== Sheet 5: Danh sách chi tiết ====================
        ws5 = wb.create_sheet("Danh sách người dùng")
        
        ws5['A1'] = "DANH SÁCH NGƯỜI DÙNG CHI TIẾT"
        ws5['A1'].font = title_font
        
        users = User.query.filter_by(is_active=True).order_by(User.created_at.desc()).limit(500).all()
        
        user_headers = ['ID', 'Username', 'Email', 'Họ tên', 'Vai trò', 'Vị trí', 'Điểm', 'Cấp độ', 'Ngày tạo']
        for col, header in enumerate(user_headers, 1):
            cell = ws5.cell(row=3, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = thin_border
        
        for row_idx, user in enumerate(users, 4):
            ws5.cell(row=row_idx, column=1, value=user.id).border = thin_border
            ws5.cell(row=row_idx, column=2, value=user.username).border = thin_border
            ws5.cell(row=row_idx, column=3, value=user.email).border = thin_border
            ws5.cell(row=row_idx, column=4, value=user.full_name or '').border = thin_border
            ws5.cell(row=row_idx, column=5, value=user.role).border = thin_border
            ws5.cell(row=row_idx, column=6, value=user.location or '').border = thin_border
            ws5.cell(row=row_idx, column=7, value=user.points or 0).border = thin_border
            ws5.cell(row=row_idx, column=8, value=user.level or 1).border = thin_border
            ws5.cell(row=row_idx, column=9, value=user.created_at.strftime('%d/%m/%Y') if user.created_at else '').border = thin_border
        
        # Auto-width for all sheets (skip merged cells)
        from openpyxl.cell.cell import MergedCell
        from openpyxl.utils import get_column_letter
        for ws in wb.worksheets:
            for col_idx in range(1, ws.max_column + 1):
                max_length = 0
                column_letter = get_column_letter(col_idx)
                for row_idx in range(1, ws.max_row + 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    if isinstance(cell, MergedCell):
                        continue
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max(max_length + 2, 10), 50)
                ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        filename = f"user_analytics_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        return send_file(
            output,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500
